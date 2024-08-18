from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import JsonResponse
from django.conf import settings
from django.http import HttpResponse
from datetime import datetime
import os
from master.forms import DepartmentForm, Designation_Form, LocationForm, EmpForm, SkillFormSet, UseraddForm ,UsereditForm
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from master.dbquery import department_list_query, designation_list_query, location_list_query, employee_list_query,user_list_query
import json
from master.models import Department,User,Designation
import openpyxl
from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from django.shortcuts import get_list_or_404
import csv
from django.contrib.auth.hashers import make_password, check_password
import logging
logger = logging.getLogger(__name__)
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponseForbidden




def indexpage(request):
    return render(request, 'index.html')

def ad_login(request):
    template_name = 'login.html'
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user_exist = User.objects.filter(username=username).exists()
        
        if user_exist:
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                if user.role == 'ADMIN' or user.role == 'VIEWER':
                    login(request, user)
                    return redirect('indexpage')
                else:
                    context = {'msg': 'Invalid Username or Password!'}
                    return render(request, template_name, context)
            else:
                context = {'msg': 'Password is incorrect!'}
                return render(request, template_name, context)
        else:
            context = {'msg': 'User Does Not exist'}
            return render(request, template_name, context)

    # Handle GET request
    else:
        return render(request, template_name)

def ad_logout(request):
    logout(request)
    return redirect(ad_login)

# def department_list(request):
#     print("ajjjjjjjjjjjh")
#     if request.method == "GET":
#         template_name = 'master/department_list.html'
#         dep = department_list_query()
#         print(dep,'dep')
#         departments = [{'department_id': item[0], 'department_name': item[1], 'description': item[2]} for item in dep]
#         context = {'departments':departments}
#         return render(request, template_name, context )
    
# @csrf_exempt
def department_list(request):
    if request.method == "GET":
        template_name = 'master/department_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        dep = department_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(dep)
    

def department_add(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    form = DepartmentForm()
    template_name = 'master/add_department.html'
    context = {'form': form}
    
    if request.method == 'POST':
        print(request.user.id,"Form submitted")
        form = DepartmentForm(request.POST, request.FILES)
        
        if form.is_valid():
            print("Form is valid")
            data = form.save()
            # data.created_by =User.objects.get(id=request.user.id)
            data.save()
           
            messages.success(request, 'Department added Successfully.', 'alert-success')
            return redirect('department_list')
            
        else:
            print("Form is not valid")
            print(form.errors)  # Print form errors to debug
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else :
        print("Rendering form")
        return render(request, template_name, context)
    
    
# @login_required(login_url='ad_login')    
def department_edit(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    template_name = 'master/department_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Department not found.', 'alert-danger')
        return redirect('department_list')
    try:
        dep_obj = Department.objects.get(department_id=pk)
    except Department.DoesNotExist:
        messages.error(request, 'Department not found.', 'alert-danger')
        return redirect('department_list')
    form = DepartmentForm(instance=dep_obj)
    context = {'form': form, 'dep_obj': dep_obj}
    if request.method == 'POST':
        form = DepartmentForm(request.POST, request.FILES, instance=dep_obj)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            messages.success(request, 'Department Successfully Updated.', 'alert-success')
            return redirect('department_list')
        else:
            print(form.errors)
            messages.error(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    
# @login_required(login_url='ad_login')
def department_detail(request,pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Invalid department ID.', 'alert-danger')
        return redirect('department_list')

    try:
        departments = Department.objects.get(department_id=pk)
    except departments.DoesNotExist:
        messages.error(request, 'Department not found.', 'alert-danger')
        return redirect('department_list')
    context = {
        'departments': departments
    }
    
    return render(request, 'master/department_detail.html', context)

# @login_required(login_url='ad_login')
def department_delete(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    department = Department.objects.get(department_id=pk)
    
    department.delete()
    messages.success(request, 'Department Deleted Successfully', 'alert-success')
    return redirect('department_list')
# def handle_uploaded_file(file):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.active

#     data = []
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         print(f"Row data: {row}")  # Log row data for debugging
#         if row[0] is not None:  # Ensure department_name is not None
#             data.append(row)

#     return data
def handle_uploaded_file(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Row data: {row}")  # Log row data for debugging
        
        # Check if row has enough columns
        if len(row) >= 6 and all(row[:6]):  # Ensure all necessary fields are present and not None
            data.append(row)
        else:
            print(f"Skipping incomplete row: {row}")  # Log skipped rows for debugging

    return data
def bulk_upload_depa(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    if request.method == 'POST':
        if 'file' not in request.FILES:
            # Handle missing file error
            return redirect('department_list')  # Or render an error template

        data = handle_uploaded_file(request.FILES['file'])
        
        for row in data:
            department_name = row[0]
            description = row[1] if len(row) > 1 else ''
            
            print(f"Processing department_name: {department_name}, description: {description}")  # Log data for debugging

            if department_name:
                if not Department.objects.filter(department_name=department_name).exists():
                    Department.objects.create(department_name=department_name, description=description)
        
        return redirect('department_list')
    
def export_departmnt(request): 
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Departments'
    columns = ['Sl.No',  'Department Name', 'Description']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title    
    for index, department in enumerate(Department.objects.all(), start=1):
        row_num += 1
        row = [
            index, 
            department.department_name,
            department.description,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departments.xlsx'

    workbook.save(response)
    return response


def download_excel(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    # Create an empty DataFrame
    df = pd.DataFrame()
    
    # Create a BytesIO object to hold the Excel file
    output = BytesIO()
    
    # Write the DataFrame to the BytesIO object
    df.to_excel(output, index=False, engine='openpyxl')
    
    # Seek to the beginning of the BytesIO object
    output.seek(0)
    
    # Create an HTTP response with the file content
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="empty_excel_sheet.xlsx"'
    
    return response
# def designation_list(request):
#     if request.method == "GET":
#         template_name = 'master/designation_list.html'
#         des = designation_list_query()
#         print(des,'des')
#         designations = [{'designation_id': item[0], 'department': item[1],'designation_name':item[2], 'description': item[3]} for item in des]
#         context = {'designations': designations}
#         return render(request, template_name, context )
def designation_list(request):
    if request.method == "GET":
        template_name = 'master/designation_list.html'
        return render(request, template_name)

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        des = designation_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(des)
    
def designation_add(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    form = Designation_Form()
    template_name = 'master/add_designation.html'
    context = {'form': form}
   
    if request.method == 'POST':
        print(request.user.id, "Form submitted")
        form = Designation_Form(request.POST, request.FILES)
        
        if form.is_valid():
            print("Form is valid")
            data = form.save()
            # try:
            #     data.created_by = User.objects.get(id=request.user.id)
            # except User.DoesNotExist:
            #     messages.error(request, 'User not found.', extra_tags='alert-danger')
            #     return redirect('designation_list')  # Or render an error template

            data.save()
            messages.success(request, 'Designation Successfully Updated.', 'alert-success')
            return redirect('designation_list')
        else:
            print("Form is not valid")
            print(form.errors)  # Print form errors to debug
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        print("Rendering form")
        return render(request, template_name, context)
    

# @login_required(login_url='adlogin')   
def designation_edit(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    template_name = 'master/designation_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Designation not found.', 'alert-danger')
        return redirect('designation_list')
    try:
        des_obj = Designation.objects.get(designation_id=pk)
    except des_obj.DoesNotExist:
        messages.error(request, 'Designation not found.', 'alert-danger')
        return redirect('designation_list')
    form = Designation_Form(instance=des_obj)
    context = {'form': form, 'des_obj': des_obj}
    if request.method == 'POST':
        form = Designation_Form(request.POST, request.FILES, instance=des_obj)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            messages.success(request, 'Designation Successfully Updated.', 'alert-success')
            return redirect('designation_list')
        else:
            print(form.errors)
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    
def designation_detail(request,pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    try:
       uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'designation not found.', 'alert-danger')
        return redirect('designation_list')

    try:
         designation = get_object_or_404(Designation,designation_id=pk)
    except designation.DoesNotExist:
        messages.error(request, 'designation not found.', 'alert-danger')
        return redirect('designation_list')
    
    department_name = designation.department.department_name
    context = {
        'designation': designation,
        'department_name':department_name
    }
    return render(request, 'master/designation_detail.html', context)

def designation_delete(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    designation = Designation.objects.get(designation_id=pk)
    designation.delete()
    messages.success(request, 'Designation Deleted Successfully', 'alert-success')
    return redirect('designation_list')
def bulk_upload_des(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    if request.method == 'POST':
       
        data = handle_uploaded_file(request.FILES['file'])
        print(data,"data")
       
        for row in data:
            department_name = row[0]
            designation_name = row[1]
            description = row[2]
            department = Department.objects.filter(department_name=department_name).first()
            if department and not Designation.objects.filter(designation_name=designation_name, department=department).exists():
                Designation.objects.create(department=department, designation_name=designation_name, description=description)

        return redirect('designation_list')
    
def export_designations_to_excel(request): 
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Departments'
    columns = ['Sl.No',  'Department Name', 'Description']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title    
    for index, department in enumerate(Department.objects.all(), start=1):
        row_num += 1
        row = [
            index,  # Sl.No
            department.department_name,
            department.description,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departments.xlsx'

    workbook.save(response)
    return response


# def location_list(request):
#     if request.method == "GET":
#         template_name = 'master/location_list.html'
#         location = location_list_query()
#         print(location,'location')
#         locations = [{'location_id': item[0], 'location_name': item[1], 'description': item[2]} for item in location]
#         context = {'location': locations}
#         return render(request, template_name, context )
def location_list(request):
    if request.method == "GET":
        template_name = 'master/location_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        loc = location_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(loc)

def locationn_add(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    form = LocationForm()
    template_name = 'master/location_add.html'
    context = {'form': form}
   
    if request.method == 'POST':
        print(request.user.id, "Form submitted")
        form = LocationForm(request.POST, request.FILES)
        
        if form.is_valid():
            print("Form is valid")
            data = form.save()
            # try:
            #     data.created_by = User.objects.get(id=request.user.id)
            # except User.DoesNotExist:
            #     messages.error(request, 'User not found.', extra_tags='alert-danger')
            #     return redirect('location_list')  # Or render an error template

            data.save()
            messages.success(request, 'Location Successfully Updated.', 'alert-success')
            return redirect('location_list')
        else:
            print("Form is not valid")
            print(form.errors)  # Print form errors to debug
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        print("Rendering form")
        return render(request, template_name, context)
    
 
def locationn_edit(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    template_name = 'master/location_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')
    try:
        loc_obj = Location.objects.get(location_id=pk)
    except loc_obj.DoesNotExist:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')
    
    form = LocationForm(instance=loc_obj)
    context = {'form': form, 'loc_obj': loc_obj}
    if request.method == 'POST':
        form = LocationForm(request.POST, request.FILES, instance=loc_obj)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            messages.success(request, 'Location Successfully Updated.', 'alert-success')
            return redirect('location_list')
        else:
            print(form.errors)
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    
def locationn_detail(request,pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')

    try:
         location = get_object_or_404(Location,location_id=pk)
    except location.DoesNotExist:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')
    context = {
        'location': location
    }
    
    return render(request, 'master/location_detail.html', context)


    

def locationn_delete(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    location = Location.objects.get(location_id=pk)
    
    location.delete()
    messages.success(request, 'Location Deleted Successfully', 'alert-success')
    return redirect('location_list')

def bulk_upload_location(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    if request.method == 'POST':
        data = handle_uploaded_file(request.FILES['file'])      
        for row in data:            
            if not Location.objects.filter(location_name=row[0]).exists():
                Location.objects.create(location_name=row[0],description=row[1])
        return redirect('location_list')
    
def export_location(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Locations'

    columns = ['Sl.No', 'Location Name', 'Description']
    row_num = 1

   
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   
    for index, location in enumerate(Location.objects.all(), start=1):
        row_num += 1
        row = [
            index,  
            location.location_name,
            location.description,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=locations.xlsx'

    workbook.save(response)
    return response

def employee_list(request):
    if request.method == "GET":
        template_name = 'master/employee_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
       
        emp = employee_list_query(start_index, page_length, search_value, draw, start_date, end_date)
       
        return JsonResponse(emp)

def employee_add(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    form = EmpForm
    formset = SkillFormSet(queryset=Skill.objects.none())
    template_name = 'master/employee_add.html'
    context = {'form': form, 'formset': formset}
   
    if request.method == 'POST':
        print(request.user.id,"Form submitted")
        form = EmpForm(request.POST, request.FILES)
        formset = SkillFormSet(request.POST, queryset=Skill.objects.none())
        if form.is_valid() and formset.is_valid():
            print("Form is valid")
            data = form.save()
            # data.created_by =User.objects.get(id=request.user.id)
            data.save()
            
            
            for skill_form in formset:
                skill = skill_form.save(commit=False)
                skill.employee = data
                skill.save()
            messages.success(request, 'Employee Added Successfully', 'alert-success')
            return redirect('employee_list')
            
        else:
            print("Form is not valid")
            print(form.errors)  
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form,'formset': formset}
            return render(request, template_name, context)
    else :
        print("Rendering form")
        return render(request, template_name, context)
def designations(request):
    department_id = request.GET.get('department')
    designations = Designation.objects.filter(department=department_id).all()
    return JsonResponse(list(designations.values('designation_id', 'designation_name')), safe=False)

@login_required(login_url='ad_login')
def employee_edit(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    template_name = 'master/employee_edit.html'
    
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Employee not found.', 'alert-danger')
        return redirect('employee_list')
    try:
        emp_obj = get_object_or_404(Employee, employee_id=pk)
    except emp_obj.DoesNotExist:
        messages.error(request, 'Employee not found.', 'alert-danger')
        return redirect('employee_list')
   
    if request.method == 'POST':
        form = EmpForm(request.POST, request.FILES, instance=emp_obj)
        formset = SkillFormSet(request.POST, queryset=Skill.objects.filter(employee=emp_obj))
        
        if form.is_valid() :
         
            employee = form.save(commit=False)
            employee.save()
           
            if formset.is_valid():
                for i in formset:
                  
                    skill = i.save()
                    skill.employee = employee
                    skill.save()
             
            messages.success(request, 'Employee Successfully Updated.', 'alert-success')
            return redirect('employee_list')
        else:
          
            messages.error(request, 'Data is not valid.', 'alert-danger')
    else:
        form = EmpForm(instance=emp_obj)
        formset = SkillFormSet(queryset=Skill.objects.filter(employee=emp_obj))
    
    context = {'form': form, 'formset': formset, 'emp_obj': emp_obj}
    return render(request, template_name, context)



def employee_detail(request,pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.") 
    try:       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'employee not found.', 'alert-danger')
        return redirect('employee_list')
    try:
         employee = get_object_or_404(Employee,employee_id=pk)
    except employee.DoesNotExist:
        messages.error(request, 'employee not found.', 'alert-danger')
        return redirect('employee_list')
    
    department=employee.department.department_name
    designation=employee.designation.designation_name
    location=employee.location.location_name
    skills = Skill.objects.filter(employee=employee)
    
    context = {
        
        'employee': employee,
        'department':department,
        'location':location,
        'designation':designation,
        'skills': skills,
    }
    
    return render(request, 'master/employee_detail.html', context)
   
def employee_delete(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    employee = Employee.objects.get(employee_id=pk)
    employee.delete()
    messages.success(request, 'Employee Deleted Successfully', 'alert-success')
    return redirect('employee_list')



def save_uploaded_file(uploaded_file, save_directory):
    try:      
        file_name, uploaded_file = next(iter(uploaded_file.items()))
        if not os.path.exists(save_directory):
            os.makedirs(save_directory)
        # Construct the full path to save the file
        save_path = os.path.join(save_directory, uploaded_file.name)
        print(save_path,'save')
        # Save the file to the specified directory
        with default_storage.open(save_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        print(f"File saved at {save_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

def bulk_upload_employee(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    if request.method == 'POST':
        if 'files' in request.FILES:
            excel_file = request.FILES['files']
            if not excel_file.name.endswith('.xlsx'):
                return HttpResponse('Invalid file format. Please upload an Excel file.', status=400)

            # Read the Excel file
            df = pd.read_excel(excel_file)

            # Process the DataFrame (assuming columns are 'Employee No', 'Name', etc.)
            for index, row in df.iterrows():
                Employee.objects.update_or_create(
                    emp_no=row['Employee No'],
                    defaults={
                        'name': row['Name'],
                        'department_name': row['Department'],
                        'designation_name': row['Designation']
                    }
                )

            return redirect('employee_list')  # Redirect to the employee list page or another page after upload

    return HttpResponse('Invalid request method.', status=405)

def filter_employees(request):
    if request.method == "GET":
        template_name = 'employee_list.html'
        return render(request, template_name)

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        emp = employee_list_query(start_index, page_length, search_value, draw, start_date, end_date)

        return JsonResponse(emp)
    

def export_employee(request):
    employee = Employee.objects.all()
    data = []

    for index, employee in enumerate(employee, start=1):
        photo_url = employee.photo.url if employee.photo and default_storage.exists(employee.photo.name) else ''

        data.append({
            'Sl.No': index,
            'Employee No': employee.emp_no,
            'Join Date': employee.join_date,
            'Name': employee.name,
            'Phone': employee.phone,
            'Address': employee.address,
            'Emp Start Date': employee.emp_start_date,
            'Emp End Date': employee.emp_end_date,
            'Photo': photo_url,
            'Status': employee.status,
            'Department': employee.department.department_name if employee.department else '',
            'Designation': employee.designation.designation_name if employee.designation else '',
            'Location': employee.location.location_name if employee.location else '',
            'Skills': ', '.join([skill.skill_name for skill in employee.skills.all()]),
            
        })

    df = pd.DataFrame(data)

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee.xlsx'

    # Write the DataFrame to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='employee')

    return response
def download_template(request):
    # Employee details template
    employee_data = {
        'Employee ID': [],
        'Join Date': [],
        'Employee No': [],
        'Name': [],
        'Phone': [],
        'Address': [],
        'Employee Start Date': [],
        'Employee End Date': [],
        'Photo': [],
        'Status': [],
        'Department': [],
        'Designation': [],
        'Location': []
    }
    
    # Skills details template
    skills_data = {
        'Skill ID': [],
        'Employee ID': [],
        'Skill Name': [],
        'Description': []
    }
    
    # Create DataFrames
    df_employee = pd.DataFrame(employee_data)
    df_skills = pd.DataFrame(skills_data)
    
    # Create a BytesIO buffer
    buffer = BytesIO()
    
    # Write the DataFrames to the buffer
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_employee.to_excel(writer, sheet_name='Employees', index=False)
        df_skills.to_excel(writer, sheet_name='Skills', index=False)
    
    # Get the content of the buffer
    buffer.seek(0)
    
    # Create the HTTP response
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=template.xlsx'
    
    return response



def download_selected(request):
    ids = request.GET.get('ids', '')
    id_list = ids.split(',')
    
    # Fetch the selected employees from the database
    employees = get_list_or_404(Employee, employee_id__in=id_list)
    
    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=selected_employees.csv'
    
    writer = csv.writer(response)
    writer.writerow(['Employee No', 'Name', 'Department', 'Designation'])
    
    for emp in employees:
        writer.writerow([emp.emp_no, emp.name, emp.department, emp.designation])
    
    return response
def generate_employee_pdf(employee_id):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    employee = Employee.objects.get(employee_id=employee_id)

    # Add a heading with a larger font
    p.setFont("Helvetica-Bold", 18)
    p.drawString(100, height - 50, "Employee Details")

 
    p.setFont("Helvetica-Bold", 12)
    p.drawString(70, height - 90, "Employee Name:")
    p.drawString(70, height - 110, "Employee Number:")
    p.drawString(70, height - 130, "Join Date:")
    p.drawString(70, height - 150, "Phone:")
    p.drawString(70, height - 170, "Address:")
    p.drawString(70, height - 190, "Start Date:")
    p.drawString(70, height - 210, "End Date:")
    p.drawString(70, height - 230, "Status:")
    p.drawString(70, height - 250, "Department:")
    p.drawString(70, height - 270, "Designation:")
    p.drawString(70, height - 290, "Location:")

    # Add employee data with regular font
    p.setFont("Helvetica", 12)
    p.drawString(200, height - 90, str(employee.name))
    p.drawString(200, height - 110, str(employee.emp_no))
    p.drawString(200, height - 130, str(employee.join_date))
    p.drawString(200, height - 150, str(employee.phone))
    p.drawString(200, height - 170, str(employee.address))
    p.drawString(200, height - 190, str(employee.emp_start_date))
    p.drawString(200, height - 210, str(employee.emp_end_date) if employee.emp_end_date else 'N/A')
    p.drawString(200, height - 230, str(employee.status))
    p.drawString(200, height - 250, str(employee.department))
    p.drawString(200, height - 270, str(employee.designation))
    p.drawString(200, height - 290, str(employee.location))

    # Add employee photo with a label
    if employee.photo:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(70, height - 320, "Employee Photo:")
        p.drawImage(employee.photo.path, 200, height - 440, width=100, preserveAspectRatio=True)
       

    # Footer
    p.setFont("Helvetica-Oblique", 10)
    p.drawString(70, 50, "Generated by datahub technologies")

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer.getvalue()


from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

def download_employee(request, employee_id):
    employee = get_object_or_404(Employee, employee_id=employee_id)
    buffer = generate_employee_pdf(employee_id)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{employee.name}_details.pdf"'
    return response
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404, render
from django.core.mail import EmailMessage
from django.http import JsonResponse
from django.template.loader import render_to_string
from io import BytesIO
from xhtml2pdf import pisa

from django.core.mail import EmailMessage, BadHeaderError
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseServerError

def generate_pdf(employee):
    # Create a PDF buffer to hold the PDF data
    buffer = BytesIO()

    # Create the PDF document with the specified page size
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Add employee photo if available
    if employee.photo:  # Assuming you have a `photo` field in your Employee model
        try:
            photo_path = employee.photo.path
            img = Image(photo_path, width=2*inch, height=2*inch)
            elements.append(img)
            elements.append(Spacer(1, 12))
        except Exception as e:
            # Handle the case where the photo cannot be loaded
            print(f"Error loading image: {e}")

    # Retrieve related data
    department = employee.department.department_name
    designation = employee.designation.designation_name
    location = employee.location.location_name
    skills = Skill.objects.filter(employee=employee)

    # Get styles for text formatting
    styles = getSampleStyleSheet()

    # Add employee details to the PDF
    elements.append(Paragraph(f"Employee No: {employee.emp_no}", styles['Title']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Name: {employee.name}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Phone: {employee.phone}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Address: {employee.address}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Employee Join Date: {employee.join_date}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Employee Start Date: {employee.emp_start_date}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Employee End Date: {employee.emp_end_date}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Status: {employee.status}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Department: {department}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Designation: {designation}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Location: {location}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    # Add skills to the PDF, handling the case where there are no skills
    if skills.exists():
        skills_list = ', '.join(skill.skill_name for skill in skills)
    else:
        skills_list = 'No skills recorded'

    elements.append(Paragraph(f"Skills: {skills_list}", styles['Normal']))

    # Build the PDF document
    doc.build(elements)

    # Get the PDF data from the buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf

def mail_pdf(request,employee_id):
     # Ensure the request method is POST
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid request method. Use POST.")

    # Get the employee object or return a 404 error if not found
    employee = get_object_or_404(Employee, employee_id=employee_id)
    
    # Generate the PDF
    pdf = generate_pdf(employee)
    
    # Get email address from POST data
    recipient_email = request.POST.get('email')
    if not recipient_email:
        return HttpResponseBadRequest("No email address provided.")
    
    # Define the email subject and body
    subject = f"Employee Report for {employee.name}"
    body = f"Dear {employee.name},\n\nPlease find attached the PDF report containing your details.\n\nBest regards,\nYour Company"
    
    try:
        # Create an email message
        email = EmailMessage(
            subject,
            body,
            'your_email@example.com',  # Replace with your sender email address
            [recipient_email],  # Replace with the employee's email address
        )
        
        # Attach the PDF
        email.attach(f"employee_{employee.emp_no}.pdf", pdf, 'application/pdf')
        
        # Send the email
        email.send()

        # Return a success response
        messages.success(request, 'Email sent successfully.','alert-success')
        return redirect('employee_list')

    except Exception as e:
        # Log the exception and return an error response
        print(f"Error sending email: {e}")
        return HttpResponseServerError("An error occurred while sending the email.")
# -------------------------------------------------user---------------

def is_admin(user):
    return user.is_authenticated and user.role == 'ADMIN'
@csrf_exempt
def user_list(request):
    if request.method == "GET":
        template_name = 'accounts/user_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
       
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        user = user_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(user)
    
@user_passes_test(is_admin)
def user_add(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    form = UseraddForm
   
    template_name = 'accounts/user_add.html'
    
    context = {'form': form}
    if request.method == 'POST':
        form = UseraddForm(request.POST, request.FILES)
        
        if form.is_valid() :
            data = form.save(commit=False)
           
            passw = data.password
            passw = make_password(passw)
            data.password = passw
           
            data.save()
            messages.success(request, 'User Added Successfully', 'alert-success')
            return redirect('user_list')
        else:
            messages.error(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form,}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    
@login_required(login_url='ad_login')
def user_edit(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    # Check if the logged-in user is a Viewer
    if request.user.role == 'VIEWER':
        messages.error(request, 'You do not have permission to edit users.', 'alert-danger')
        return redirect('user_list')

    template_name = 'accounts/user_edit.html'

    try:
        # Fetch the user by primary key (assumed to be an integer)
        user_obj = User.objects.get(pk=pk)
    except User.DoesNotExist:
        messages.error(request, 'User not found.', 'alert-danger')
        return redirect('user_list')

    if request.method == 'POST':
        form = UsereditForm(request.POST, request.FILES, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'User Updated Successfully', 'alert-success')
            return redirect('user_list')
        else:
            messages.error(request, 'Data is not valid.', 'alert-danger')
    else:
        form = UsereditForm(instance=user_obj)

    context = {'form': form}
    return render(request, template_name, context)
    
def user_detail(request,pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    try:
         user = get_object_or_404(User,id=pk)
    except user.DoesNotExist:
        messages.error(request, 'user not found.', 'alert-danger')
        return redirect('user_list')    
    context = {
        'user': user
    }    
    return render(request, 'accounts/user_detail.html', context)

@login_required(login_url='ad_login')
@user_passes_test(is_admin)
def user_delete(request, pk):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    user = User.objects.get(id=pk)    
    user.delete()
    messages.success(request, 'User Deleted Successfully', 'alert-success')
    return redirect('user_list')
def bulk_upload_user(request):
    if request.user.role == 'VIEWER':
        return HttpResponseForbidden("You are not allowed to perform this action.")
    if request.method == 'POST':
        data = handle_uploaded_file(request.FILES['file'])
      
        for row in data:
            try:
                username = row[0]
                email = row[1]
                first_name = row[2]
                last_name = row[3]
                password = row[4]
                role = row[5]
                
                # Check for existing username
                if not User.objects.filter(username=username).exists():
                    User.objects.create(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        role=role,
                        password=make_password(password)
                    )
                else:
                    logging.warning(f"Username {username} already exists. Skipping this entry.")
            except IndexError:
                logging.error(f"Row is incomplete or malformed: {row}")
                
        return redirect('user_list')
    
    

def export_user(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    columns = ['Sl.No', 'Username', 'Email', 'First Name', 'Last Name']
    row_num = 1

   
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for index, user in enumerate(User.objects.all(), start=1):
        row_num += 1
        row = [
            index,
            user.username,
            user.email,
            user.first_name,
            user.last_name,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'

    workbook.save(response)
    return response



